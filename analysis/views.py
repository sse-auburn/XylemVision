import json
import zipfile
import numpy as np
import cv2
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from PIL import Image
from io import BytesIO
import base64
import openpyxl
from .engine import progressive_yolo_sam, sam_predictor, calculate_metrics
from .utils import refine_masks, blend_mask, draw_boxes, class_color_cycle, mask_from_contour, compute_props

_last_analysis_cache = {}


def pil_to_base64(img):
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    encoded = base64.b64encode(buffer.getvalue()).decode()
    return f"data:image/png;base64,{encoded}"


def root_analysis_view(request):
    return render(request, 'upload.html', {})


# ─── SSE streaming analysis endpoint ─────────────────────────────────────────

def analyze_stream_view(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("POST required")

    images = request.FILES.getlist('image')
    if not images:
        return HttpResponseBadRequest("No images uploaded")

    from django.http import StreamingHttpResponse

    def event_stream():
        total = len(images)
        for i, image_file in enumerate(images):
            yield f"data: {json.dumps({'type':'start','file':image_file.name,'index':i,'total':total})}\n\n"
            try:
                image = Image.open(image_file).convert('RGB')
                result, original_img, overlay_img = progressive_yolo_sam(image)

                xylem_details = result['metrics'].get('xylem_details', [])
                colours = [c for c in result.get('colours', []) if c.get('class') == 'Xylem']
                merged = []
                for j in range(max(len(xylem_details), len(colours))):
                    detail = xylem_details[j] if j < len(xylem_details) else None
                    color  = colours[j]        if j < len(colours)        else None
                    if detail and color:
                        merged.append({'instance': j, **detail, 'rgb': list(color['rgb'])})
                    elif detail:
                        merged.append({'instance': j, **detail})

                colour_map = {}
                for c in result.get('colours', []):
                    colour_map.setdefault(c['class'], []).append(list(c['rgb']))

                payload = {
                    'type':           'result',
                    'file':           image_file.name,
                    'original_image': pil_to_base64(original_img),
                    'overlay_image':  pil_to_base64(overlay_img),
                    'n_xylem':        result['n_xylem'],
                    'n_vb':           result['n_vb'],
                    'n_root':         result['n_root'],
                    'metrics':        result['metrics'],
                    'merged_xylem':   merged,
                    'contours':       result['contours'],
                    'colour_map':     colour_map,
                    'img_width':      original_img.width,
                    'img_height':     original_img.height,
                }
                _last_analysis_cache[image_file.name] = {
                    **payload,
                    'rgb_np': result['rgb_np'],
                    'scale':  result.get('scale', 1.0),
                    'boxes':  result['boxes'],
                    'labels': result['labels'],
                }
                yield f"data: {json.dumps(payload)}\n\n"
            except Exception as e:
                yield f"data: {json.dumps({'type':'error','file':image_file.name,'message':str(e)})}\n\n"

        yield f"data: {json.dumps({'type':'done','total':total})}\n\n"

    resp = StreamingHttpResponse(event_stream(), content_type='text/event-stream')
    resp['Cache-Control']      = 'no-cache'
    resp['X-Accel-Buffering']  = 'no'
    return resp


# ─── SAM Prompt endpoint ──────────────────────────────────────────────────────

def sam_prompt_view(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("POST required")

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON")

    filename = body.get('filename')
    if not filename or filename not in _last_analysis_cache:
        return HttpResponseBadRequest("Image not found in cache")

    cached = _last_analysis_cache[filename]
    rgb_np = cached['rgb_np']
    scale  = cached.get('scale', 1.0)
    H, W = rgb_np.shape[:2]
    mode = body.get('mode', 'box')

    sam_predictor.set_image(rgb_np)

    if mode == 'box':
        box = body.get('box')
        if not box or len(box) != 4:
            return HttpResponseBadRequest("box must be [x1,y1,x2,y2]")
        # Frontend coords are in original image space — scale down to processed space
        box_np = np.array([v * scale for v in box], dtype=float)
        masks, scores, _ = sam_predictor.predict(
            box=box_np, multimask_output=True
        )
    elif mode == 'point':
        pt = body.get('point')
        if not pt or len(pt) != 2:
            return HttpResponseBadRequest("point must be [x,y]")
        coords = np.array([[pt[0] * scale, pt[1] * scale]], dtype=float)
        labels = np.array([1])
        masks, scores, _ = sam_predictor.predict(
            point_coords=coords, point_labels=labels, multimask_output=True
        )
    else:
        return HttpResponseBadRequest("mode must be 'box' or 'point'")

    scores = scores.squeeze()
    best_mask = masks[int(scores.argmax())]

    # Extract largest contour from the mask
    cnts, _ = cv2.findContours(
        best_mask.astype(np.uint8), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
    )
    if not cnts:
        return JsonResponse({'contour': []})

    cnt     = max(cnts, key=cv2.contourArea)
    raw     = cnt.reshape(-1, 2).tolist()
    # Scale contour back to original image coordinates
    contour = [[round(x / scale), round(y / scale)] for x, y in raw] if scale != 1.0 else raw
    return JsonResponse({'contour': contour})


# ─── Reanalyze endpoint ───────────────────────────────────────────────────────

def reanalyze_view(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("POST required")

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON")

    filename = body.get('filename')
    if not filename or filename not in _last_analysis_cache:
        return HttpResponseBadRequest("Image not found in cache")

    cached = _last_analysis_cache[filename]
    rgb_np = cached['rgb_np']
    H, W   = rgb_np.shape[:2]
    scale  = cached.get('scale', 1.0)
    s2     = scale * scale

    polygons = body.get('polygons', [])  # [{class, points:[[x,y],...]}]

    # Points arrive in original-image space; scale down to rgb_np space for mask creation
    def _to_model(pts):
        if scale == 1.0:
            return pts
        return [[round(x * scale), round(y * scale)] for x, y in pts]

    # Group polygon points by class
    cls2pts = {'Xylem': [], 'Vascular bundle': [], 'Total root': []}
    for p in polygons:
        cls = p.get('class')
        pts = p.get('points', [])
        if cls in cls2pts and pts:
            cls2pts[cls].append(_to_model(pts))

    # Rebuild binary masks from polygon contours
    x_masks  = [mask_from_contour(pts, H, W) for pts in cls2pts['Xylem']]
    vb_masks = [mask_from_contour(pts, H, W) for pts in cls2pts['Vascular bundle']]
    rt_masks = [mask_from_contour(pts, H, W) for pts in cls2pts['Total root']]

    # Refine (subtract overlaps) and compute metrics
    x_masks, vb_masks, rt_masks = refine_masks(x_masks, vb_masks, rt_masks)
    metrics = calculate_metrics(x_masks, vb_masks, rt_masks)

    # Scale metrics back to original-image pixel space
    if scale != 1.0:
        metrics['vb_total_area']     /= s2
        metrics['vb_max_diameter']   /= scale
        metrics['root_total_area']   /= s2
        metrics['root_max_diameter'] /= scale
        for d in metrics['xylem_details']:
            d['area']     /= s2
            d['diameter'] /= scale

    # Extract contours and scale back to original-image space
    def _to_orig(contour):
        if scale == 1.0:
            return contour
        return [[round(x / scale), round(y / scale)] for x, y in contour]

    x_contours  = [_to_orig(p['contour']) for p in compute_props(x_masks,  return_contours=True)]
    vb_contours = [_to_orig(p['contour']) for p in compute_props(vb_masks, return_contours=True)]
    rt_contours = [_to_orig(p['contour']) for p in compute_props(rt_masks, return_contours=True)]

    # Rebuild overlay image (in rgb_np space — no upscaling needed for display)
    overlay = rgb_np.copy()
    colour_meta = []
    colour_map = {}
    for cls, masks in [('Xylem', x_masks), ('Vascular bundle', vb_masks), ('Total root', rt_masks)]:
        for k, m in enumerate(masks):
            col = class_color_cycle(cls, k)
            overlay = blend_mask(overlay, m, col)
            colour_meta.append({'class': cls, 'inst': k, 'rgb': list(col)})
            colour_map.setdefault(cls, []).append(list(col))

    # Derive bounding boxes from current corrected masks (rgb_np space) — not the stale cached YOLO boxes
    new_boxes, new_labels = [], []
    for cls_name, masks in [('Xylem', x_masks), ('Vascular bundle', vb_masks), ('Total root', rt_masks)]:
        for m in masks:
            ys, xs = np.where(m)
            if len(xs):
                new_boxes.append([float(xs.min()), float(ys.min()), float(xs.max()), float(ys.max())])
                new_labels.append(cls_name)
    overlay = draw_boxes(overlay, np.array(new_boxes) if new_boxes else np.empty((0, 4)), new_labels)
    overlay_pil = Image.fromarray(overlay)
    overlay_b64 = pil_to_base64(overlay_pil)

    # Rebuild merged_xylem
    xylem_details = metrics.get('xylem_details', [])
    x_colours = [c for c in colour_meta if c['class'] == 'Xylem']
    merged = []
    for i in range(max(len(xylem_details), len(x_colours))):
        detail = xylem_details[i] if i < len(xylem_details) else None
        color  = x_colours[i]    if i < len(x_colours)    else None
        if detail and color:
            merged.append({'instance': i, **detail, 'rgb': list(color['rgb'])})
        elif detail:
            merged.append({'instance': i, **detail})

    # Update cache
    cached['overlay_image'] = overlay_b64
    cached['n_xylem']       = len(x_masks)
    cached['n_vb']          = len(vb_masks)
    cached['n_root']        = len(rt_masks)
    cached['metrics']       = metrics
    cached['merged_xylem']  = merged
    cached['contours']      = {'Xylem': x_contours, 'Vascular bundle': vb_contours, 'Total root': rt_contours}
    cached['colour_map']    = colour_map

    return JsonResponse({
        'overlay_image': overlay_b64,
        'n_xylem':  len(x_masks),
        'n_vb':     len(vb_masks),
        'n_root':   len(rt_masks),
        'metrics':  metrics,
        'merged_xylem': merged,
        'contours': {'Xylem': x_contours, 'Vascular bundle': vb_contours, 'Total root': rt_contours},
        'colour_map': colour_map,
    })


# ─── Merge Masks endpoint ────────────────────────────────────────────────────

def merge_masks_view(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("POST required")
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON")

    filename = body.get('filename')
    if not filename or filename not in _last_analysis_cache:
        return HttpResponseBadRequest("Image not found in cache")

    cached = _last_analysis_cache[filename]
    rgb_np = cached['rgb_np']
    H, W = rgb_np.shape[:2]

    polygons = body.get('polygons', [])  # [{points: [[x,y],...]}]
    merged_mask = np.zeros((H, W), dtype=bool)
    for pg in polygons:
        pts = pg.get('points', [])
        if pts:
            merged_mask |= mask_from_contour(pts, H, W)

    cnts, _ = cv2.findContours(merged_mask.astype(np.uint8), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not cnts:
        return JsonResponse({'contour': []})

    cnt = max(cnts, key=cv2.contourArea)
    return JsonResponse({'contour': cnt.reshape(-1, 2).tolist()})


# ─── Export training data (image + CVAT XML) as ZIP ─────────────────────────

def _build_cvat_xml(filename, img_w, img_h, polygons):
    """Build CVAT-format XML bytes for one image."""
    lines = [
        '<annotation>',
        f'  <filename>{filename}</filename>',
        '  <folder></folder>',
        '  <source>',
        '    <sourceImage></sourceImage>',
        '    <sourceAnnotation>XylemVision</sourceAnnotation>',
        '  </source>',
        '  <imagesize>',
        f'    <nrows>{img_h}</nrows>',
        f'    <ncols>{img_w}</ncols>',
        '  </imagesize>',
    ]
    for obj_id, poly in enumerate(polygons):
        cls = poly.get('cls', '')
        pts = poly.get('points', [])
        lines += [
            '  <object>',
            f'    <name>{cls}</name>',
            '    <deleted>0</deleted>',
            '    <verified>0</verified>',
            '    <occluded>no</occluded>',
            '    <date></date>',
            f'    <id>{obj_id}</id>',
            '    <parts>',
            '      <hasparts></hasparts>',
            '      <ispartof></ispartof>',
            '    </parts>',
            '    <polygon>',
        ]
        for pt in pts:
            lines += [
                '      <pt>',
                f'        <x>{pt[0]:.2f}</x>',
                f'        <y>{pt[1]:.2f}</y>',
                '      </pt>',
            ]
        lines += [
            '      <username></username>',
            '    </polygon>',
            '    <attributes></attributes>',
            '  </object>',
        ]
    lines.append('</annotation>')
    return '\n'.join(lines).encode('utf-8')


def export_training_view(request):
    """Export a single image + its canvas polygons as a training ZIP."""
    if request.method != 'POST':
        return HttpResponseBadRequest("POST required")
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON")

    filename = body.get('filename')
    if not filename or filename not in _last_analysis_cache:
        return HttpResponseBadRequest("Image not found in cache")
    polygons = body.get('polygons', [])
    if not polygons:
        return HttpResponseBadRequest("No polygons provided")

    cached = _last_analysis_cache[filename]
    stem   = filename.rsplit('.', 1)[0] if '.' in filename else filename
    xml_bytes = _build_cvat_xml(filename, cached.get('img_width', 0),
                                cached.get('img_height', 0), polygons)

    orig_b64 = cached.get('original_image', '')
    if ',' not in orig_b64:
        return HttpResponseBadRequest("Original image not available in cache")
    img_bytes = base64.b64decode(orig_b64.split(',', 1)[1])

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'images/{filename}', img_bytes)
        zf.writestr(f'labels/{stem}.xml', xml_bytes)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{stem}_training.zip"'
    return response


def export_training_batch_view(request):
    """Export multiple checked images + their canvas polygons as one ZIP."""
    if request.method != 'POST':
        return HttpResponseBadRequest("POST required")
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponseBadRequest("Invalid JSON")

    images = body.get('images', [])  # [{filename, polygons:[{cls,points},...]}]
    if not images:
        return HttpResponseBadRequest("No images provided")

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for item in images:
            filename = item.get('filename', '')
            polygons = item.get('polygons', [])
            if not filename or filename not in _last_analysis_cache or not polygons:
                continue
            cached = _last_analysis_cache[filename]
            stem   = filename.rsplit('.', 1)[0] if '.' in filename else filename
            xml_bytes = _build_cvat_xml(filename, cached.get('img_width', 0),
                                        cached.get('img_height', 0), polygons)
            orig_b64 = cached.get('original_image', '')
            if ',' not in orig_b64:
                continue
            img_bytes = base64.b64decode(orig_b64.split(',', 1)[1])
            zf.writestr(f'images/{filename}', img_bytes)
            zf.writestr(f'labels/{stem}.xml', xml_bytes)

    buf.seek(0)
    response = HttpResponse(buf, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="training_data.zip"'
    return response


# ─── Download all overlay images as ZIP ──────────────────────────────────────

def download_overlays_view(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("POST required")

    filenames = request.POST.getlist('filenames')
    if not filenames:
        return HttpResponseBadRequest("No files selected")

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for filename in filenames:
            if filename not in _last_analysis_cache:
                continue
            cached = _last_analysis_cache[filename]
            # overlay_image is a data URI: "data:image/png;base64,<b64>"
            b64 = cached['overlay_image'].split(',', 1)[1]
            img_bytes = base64.b64decode(b64)
            stem = filename.rsplit('.', 1)[0]
            zf.writestr(f"{stem}_overlay.png", img_bytes)

    buf.seek(0)
    response = HttpResponse(buf, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="overlays.zip"'
    return response


# ─── Helper: create XLSX workbook for one analysis result ─────────────────────
# Format matches pipeline2_final.py: Summary sheet + Xylem Details sheet

def generate_xlsx_for_result(result):
    wb = openpyxl.Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Image", "Xylem Count", "Vascular Bundle Area",
                   "Vascular Bundle Diameter", "Total Root Diameter"])
    ws_sum.append([
        result['file'],
        result['n_xylem'],
        result['metrics'].get('vb_total_area', 0),
        result['metrics'].get('vb_max_diameter', 0),
        result['metrics'].get('root_max_diameter', 0),
    ])

    ws_xy = wb.create_sheet(title="Xylem Details")
    ws_xy.append(["Image", "Xylem ID", "Xylem Area", "Xylem Diameter"])
    for d in result['metrics'].get('xylem_details', []):
        ws_xy.append([result['file'], f"x_{d['id'] + 1}", d['area'], d['diameter']])

    return wb


# ─── Download XLSX for a single analysis ─────────────────────────────────────

def download_xlsx(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("Invalid request method")

    filename = request.POST.get('filename')
    if not filename or filename not in _last_analysis_cache:
        return HttpResponseBadRequest("No analysis found for that file")

    result = _last_analysis_cache[filename]
    wb = generate_xlsx_for_result(result)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}_analysis.xlsx"'
    wb.save(response)
    return response


# ─── Download XLSX for all analyses combined ─────────────────────────────────
# Two-sheet format matching pipeline2_final.py: Summary (one row/image) + Xylem Details

def download_all_xlsx(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("Invalid request method")

    filenames = request.POST.getlist('filenames')
    if not filenames:
        return HttpResponseBadRequest("No files selected")

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Image", "Xylem Count", "Vascular Bundle Area",
                   "Vascular Bundle Diameter", "Total Root Diameter"])

    ws_xy = wb.create_sheet(title="Xylem Details")
    ws_xy.append(["Image", "Xylem ID", "Xylem Area", "Xylem Diameter"])

    for filename in filenames:
        if filename not in _last_analysis_cache:
            continue
        result = _last_analysis_cache[filename]
        ws_sum.append([
            filename,
            result['n_xylem'],
            result['metrics'].get('vb_total_area', 0),
            result['metrics'].get('vb_max_diameter', 0),
            result['metrics'].get('root_max_diameter', 0),
        ])
        for d in result['metrics'].get('xylem_details', []):
            ws_xy.append([filename, f"x_{d['id'] + 1}", d['area'], d['diameter']])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="all_analysis_combined.xlsx"'
    wb.save(response)
    return response
